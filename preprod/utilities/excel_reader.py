import os
import openpyxl

def read_test_data():
    file_path = "C://Users//gaurav//Desktop//iFOLIO//Automation//iFOLIO_auto//utilities//test_data.xlsx"

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    email = sheet.cell(row=2, column=1).value
    password = sheet.cell(row=2, column=2).value
    user = sheet.cell(row=2,column=3).value
    ifolio_no = sheet.cell(row=2,column=4).value
    single_share_name = sheet.cell(row=2,column=5).value
    single_share_company = sheet.cell(row=2,column=6).value
    single_share_email = sheet.cell(row=2,column=7).value
    email_message = sheet.cell(row=2,column=8).value

    user_email = sheet.cell(row=5, column=1).value
    user_password = sheet.cell(row=5, column=2).value

    campaign_name = sheet.cell(row=8, column=1).value
    group_message = sheet.cell(row=8, column=2).value

    return {'email': email, 'password': password,'user':user,'iFOLIO No':ifolio_no,'Single_Share_Name':single_share_name,
            'Single_Share_Company':single_share_company,'Single_Share_Email':single_share_email,
            'user_email':user_email,'user_password':user_password,'Email_message':email_message,
            'Campaign_name':campaign_name,'Group_message':group_message}

def read_beta_test_data():
    file_path = "C://Users//gaurav//Desktop//iFOLIO//Automation//iFOLIO_auto//utilities//Beta_test_data.xlsx"

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    #Add account details
    Login_email = sheet.cell(row=11, column=1).value
    Login_password = sheet.cell(row=11, column=2).value
    Plan_type = sheet.cell(row=2, column=1).value
    Account_name = sheet.cell(row=2, column=2).value
    First_name = sheet.cell(row=2,column=3).value
    Last_name = sheet.cell(row=2,column=4).value
    Email = sheet.cell(row=2,column=5).value
    Phone = sheet.cell(row=2,column=6).value
    Term = sheet.cell(row=2,column=7).value
    iFOLIO_owner_name = sheet.cell(row=2,column=8).value
    iFOLIO_owner_email = sheet.cell(row=2,column=9).value
    Renewal_date = sheet.cell(row=2,column=10).value
    Number_of_managers = sheet.cell(row=2,column=11).value
    Number_of_users = sheet.cell(row=2,column=12).value
    Account_email_bank_value = sheet.cell(row=2, column=13).value
    User_email_bank_value = sheet.cell(row=2, column=14).value
    Overage_email_price = sheet.cell(row=2, column=15).value
    Account_text_bank_value = sheet.cell(row=2, column=16).value
    User_text_bank_value = sheet.cell(row=2, column=17).value
    Overage_text_price = sheet.cell(row=2, column=18).value
    Site_limit_per_account = sheet.cell(row=2, column=19).value
    Site_limit_per_user = sheet.cell(row=2, column=20).value
    Contact_limit_per_account = sheet.cell(row=2, column=21).value
    Contact_limit_per_user = sheet.cell(row=2, column=22).value
    Media_library_org = sheet.cell(row=2, column=23).value
    Media_library_myupload = sheet.cell(row=2, column=24).value
    Services = sheet.cell(row=2, column=25).value

    Search_account_name = sheet.cell(row=5,column=1).value
    Search_first_name = sheet.cell(row=5,column=2).value
    Search_last_name = sheet.cell(row=5, column=3).value
    Search_email = sheet.cell(row=5,column=4).value

    Delete_email = sheet.cell(row=8,column=1).value

    #Add User Details
    Select_Account = sheet.cell(row=15,column=1).value
    User_First_Name = sheet.cell(row=15,column=2).value
    User_Last_Name = sheet.cell(row=15,column=3).value
    User_Email = sheet.cell(row=15,column=4).value
    User_Phone = sheet.cell(row=15, column=5).value
    User_Address = sheet.cell(row=15, column=6).value
    User_Job = sheet.cell(row=15, column=7).value
    User_Profile = sheet.cell(row=15, column=8).value

    #Search User Details
    US_First_Name = sheet.cell(row=19, column=1).value
    US_Last_Name = sheet.cell(row=19, column=2).value

    #For creating new template in admin panel
    Select_ifolio = sheet.cell(row=22, column=1).value
    Rename_ifolio = sheet.cell(row=22,column=2).value

    return {'login_email':Login_email,'login_password':Login_password,'plan_type': Plan_type, 'account_name': Account_name,'first_name':First_name,'last_name':Last_name,
            'email':Email,'phone':Phone,'term':Term,'iFOLIO_owner_name':iFOLIO_owner_name,'iFOLIO_owner_email':iFOLIO_owner_email,
            'renewal_date':Renewal_date,'number_of_managers':Number_of_managers,'number_of_users':Number_of_users,
            'account_email_bank_value':Account_email_bank_value,'user_email_bank_value':User_email_bank_value,
            'overage_email_price':Overage_email_price,'account_text_bank_value':Account_text_bank_value,
            'user_text_bank_value':User_text_bank_value,'overage_text_price':Overage_text_price,'site_limit_per_account':Site_limit_per_account,
            'site_limit_per_user':Site_limit_per_user,'contact_limit_per_account':Contact_limit_per_account,
            'contact_limit_per_user':Contact_limit_per_user,'media_library_org':Media_library_org,'media_library_myupload':Media_library_myupload,
            'services':Services,'search_Account_name':Search_account_name,'search_First_name':Search_first_name,'search_Last_name':Search_last_name,
            'search_Email':Search_email,'delete_email':Delete_email,'select_Account':Select_Account,'user_First_Name':User_First_Name,
            'user_Last_Name':User_Last_Name,'user_Email':User_Email,'user_Phone':User_Phone,'user_Address':User_Address,
            'user_Job':User_Job,'user_Profile':User_Profile,'us_First_Name':US_First_Name,'us_Last_Name':US_Last_Name,
            'select_ifolio':Select_ifolio,'rename_ifolio':Rename_ifolio}

def read_preprod_test_data():
    file_path = "C://Users//gaurav//Desktop//iFOLIO//Automation//iFOLIO_auto//utilities//preprod_test_data.xlsx"

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    email = sheet.cell(row=2, column=1).value
    password = sheet.cell(row=2, column=2).value
    user = sheet.cell(row=2,column=3).value
    ifolio_no = sheet.cell(row=2,column=4).value
    single_share_name = sheet.cell(row=2,column=5).value
    single_share_company = sheet.cell(row=2,column=6).value
    single_share_email = sheet.cell(row=2,column=7).value
    email_message = sheet.cell(row=2,column=8).value

    user_email = sheet.cell(row=5, column=1).value
    user_password = sheet.cell(row=5, column=2).value

    campaign_name = sheet.cell(row=8, column=1).value
    group_message = sheet.cell(row=8, column=2).value

    return {'email': email, 'password': password,'user':user,'iFOLIO No':ifolio_no,'Single_Share_Name':single_share_name,
            'Single_Share_Company':single_share_company,'Single_Share_Email':single_share_email,
            'user_email':user_email,'user_password':user_password,'Email_message':email_message,
            'Campaign_name':campaign_name,'Group_message':group_message}





















